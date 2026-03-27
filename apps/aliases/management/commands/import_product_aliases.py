import sys

from django.core.management.base import BaseCommand

from apps.aliases.models import GlobalProductAlias, VALID_CATEGORIES


class Command(BaseCommand):
    help = 'Import product aliases from the categorized_products.xlsx Excel file.'

    def add_arguments(self, parser):
        parser.add_argument('--file', required=True, help='Path to the .xlsx file')
        parser.add_argument('--batch-size', type=int, default=5000, help='bulk_create batch size')
        parser.add_argument('--dry-run', action='store_true', help='Print stats without inserting')

    def handle(self, *args, **options):
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.stderr.write(self.style.ERROR(
                'openpyxl is required. Install it with: pip install openpyxl'
            ))
            sys.exit(1)

        file_path = options['file']
        batch_size = options['batch_size']
        dry_run = options['dry_run']

        self.stdout.write(f'Loading workbook from {file_path} (read-only mode)...')
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        # Column layout (0-indexed):
        # A=Display Name, B=Category, C-G=Alias 1-5,
        # H-L=Alias 15chr #1-5, M-Q=Alias 20chr #1-5, R-V=Alias 24chr #1-5
        ALIAS_COLS = list(range(2, 22))  # columns C through V (indices 2-21)

        rows_to_create = []
        products_processed = 0
        skipped_no_name = 0
        skipped_bad_category = 0
        valid_cats_upper = {c.upper() for c in VALID_CATEGORIES}

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                skipped_no_name += 1
                continue

            display_name = str(row[0]).strip()
            raw_category = str(row[1]).strip().upper() if row[1] else ''

            if raw_category not in valid_cats_upper:
                skipped_bad_category += 1
                continue

            category = raw_category
            products_processed += 1

            # Collect unique alias texts for this product (including the display name itself).
            seen_upper = set()
            alias_texts = []

            # Add the canonical display name as an alias too.
            dn_upper = display_name.strip().upper()
            if dn_upper and dn_upper not in seen_upper:
                seen_upper.add(dn_upper)
                alias_texts.append(display_name)

            for col_idx in ALIAS_COLS:
                if col_idx < len(row) and row[col_idx]:
                    val = str(row[col_idx]).strip()
                    val_upper = val.upper()
                    if val_upper and val_upper not in seen_upper:
                        seen_upper.add(val_upper)
                        alias_texts.append(val)

            for alias in alias_texts:
                rows_to_create.append(GlobalProductAlias(
                    canonical_name=display_name,
                    alias_text=alias,
                    alias_text_upper=alias.strip().upper(),
                    category=category,
                    source='seed',
                ))

            if products_processed % 10000 == 0:
                self.stdout.write(f'  processed {products_processed} products ...')

        wb.close()

        self.stdout.write(f'\nProducts processed: {products_processed}')
        self.stdout.write(f'Alias rows to create: {len(rows_to_create)}')
        self.stdout.write(f'Skipped (no display name): {skipped_no_name}')
        self.stdout.write(f'Skipped (bad/missing category): {skipped_bad_category}')

        if dry_run:
            self.stdout.write(self.style.WARNING('\n--dry-run: no rows inserted.'))
            return

        self.stdout.write(f'\nInserting in batches of {batch_size} (ignore_conflicts=True)...')
        created = 0
        for i in range(0, len(rows_to_create), batch_size):
            batch = rows_to_create[i:i + batch_size]
            objs = GlobalProductAlias.objects.bulk_create(batch, batch_size=batch_size, ignore_conflicts=True)
            created += len(objs)
            self.stdout.write(f'  batch {i // batch_size + 1}: {len(objs)} rows')

        self.stdout.write(self.style.SUCCESS(
            f'\nDone. {created} alias rows inserted ({len(rows_to_create) - created} conflicts skipped).'
        ))
